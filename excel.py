from time import sleep
import openpyxl
import datetime

class Excel:
    def companies_name(self):
        wb = openpyxl.load_workbook("Master.xlsx")
        sheet = wb.worksheets[0]

        i = 3
        cell_ = sheet.cell(row = i , column = 2 ).value
        companies = []
        while cell_ is not None:
            cell_ = sheet.cell(row = i , column = 2 ).value
            if cell_ is None:
                break
            companies.append(cell_)
            i += 1
        return companies